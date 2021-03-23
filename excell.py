# Author: Jeshwanth Kumar Ega(99003737)
# Contact: jeshwanth.ega@ltts.com /jeshwanthega0@gmail.com
# Date of creation: 22/3/2021
# -------------------------------------------------------------------------------------------------------------#
# -------------------------------------------------------------------------------------------------------------#


import openpyxl
# creating work bench which is noting but excile file
from openpyxl import Workbook

"""
This program performs the task of putting
all the data of a particular candidate from 5 sheets
to a master sheet provided the user gives the input of name ,
ps no or email id of that particular candidate.
excel files should be placed in different directories (minimum 4 directory)
and path access from code should not be hard
coded; master file should get the data from all excel files
This programs uses openpyxl library and it has
object oriented programming concepts such as class,
object and function.
"""
# -------------------------------------------------------------------------------------------------------------#
# -------------------------------------------------------------------------------------------------------------#
"""
This program uses openpyxl library and import openpyxl
imports the library and rather than using the name openpyxl ,
it's instructed to use the name pd instead.
From  ExcelWriter is imported in order to
write the header in the first master sheet.
"""

excel_file = Workbook()
# taking no. of files input from user
num_of_files = int(input("Enter Number of file:"))
# intiliazing empty list to store files
files = []
print("Enter file directories: ")
# taking input directories from user and storing in to the files list
for i in range(num_of_files):
    files.append(input())     # adding in to the files list
    print(files)
# creating in a new sheet already created in a work bench
excel_sheet = excel_file.create_sheet(title='MasterSheet11', index=0)

num = int(input("Enter number of persons: "))
# taking no. of persons to be serached and looping overthem
for n in range(1, num+1):
    a_in = int(input("enter ps number: "))
    b_in = input("enter name: ")
    c_in = input("enter mailid: ")
    m = 1

    for i in files:
        file = openpyxl.load_workbook(i)
        sheets = file.sheetnames
        sh = file[sheets[0]]
        maximum_row = sh.max_row
        maximum_column = sh.max_column

        if m <= 10:
            for r in range(1, maximum_row + 1):
                if sh.cell(row=r, column=1).value == a_in and sh.cell(
                        row=r, column=2).value == b_in and sh.cell(row=r,
                                                column=3).value == c_in:
                    for c in range(1, maximum_column+1):
                        if n == 1:
                            str1 = 'A' + str(m)
                            str2 = 'B' + str(m)
                            m = m + 1
                            excel_sheet[str1] = str(sh.cell(row=1,
                                                            column=c).value)
                            excel_sheet[str2] = sh.cell(row=r, column=c).value
                        else:
                            str1 = 'E' + str(m)
                            str2 = 'F' + str(m)
                            m = m + 1
                            excel_sheet[str1] = str(sh.cell(row=1,
                                                            column=c).value)
                            excel_sheet[str2] = sh.cell(row=r, column=c).value
        else:
            for r in range(4, maximum_row + 1):
                if sh.cell(row=r, column=1).value == a_in and sh.cell(
                        row=r, column=2).value == b_in and sh.cell(row=r,
                                                column=3).value == c_in:
                    for c in range(4, maximum_column + 1):
                        if n == 1:
                            str1 = 'A' + str(m)
                            str2 = 'B' + str(m)
                            m = m + 1
                            excel_sheet[str1] = str(sh.cell(row=1,
                                                            column=c).value)
                            excel_sheet[str2] = sh.cell(row=r, column=c).value
                        else:
                            str1 = 'E' + str(m)
                            str2 = 'F' + str(m)
                            m = m + 1
                            excel_sheet[str1] = str(sh.cell(row=1,
                                                            column=c).value)
                            excel_sheet[str2] = sh.cell(row=r, column=c).value
    output_dir = input("Enter output directory: ")
    out_file = input("Enter output file name: ")
    out_file_dir = output_dir + "\\" + out_file
    excel_file.save(filename=out_file_dir)
    print("File Created")
